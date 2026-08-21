#!/usr/bin/env python3
"""Audit a PPTX package and print a compact JSON structural summary."""

from __future__ import annotations

import argparse
import datetime as dt
import io
import json
import posixpath
import re
import sys
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET


NS = {
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "p": "http://schemas.openxmlformats.org/presentationml/2006/main",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "c": "http://schemas.openxmlformats.org/drawingml/2006/chart",
}


def read_text(zf: zipfile.ZipFile, name: str) -> str:
    try:
        return zf.read(name).decode("utf-8", errors="replace")
    except KeyError:
        return ""


def parse_xml(text: str):
    if not text:
        return None
    try:
        return ET.fromstring(text)
    except ET.ParseError:
        return None


def natural_key(name: str) -> tuple:
    return tuple(int(part) if part.isdigit() else part for part in re.split(r"(\d+)", name))


def local_rels_path(part_name: str) -> str:
    parent, filename = part_name.rsplit("/", 1)
    return f"{parent}/_rels/{filename}.rels"


def rel_targets(zf: zipfile.ZipFile, part_name: str) -> list[dict[str, str]]:
    rels = parse_xml(read_text(zf, local_rels_path(part_name)))
    out: list[dict[str, str]] = []
    if rels is None:
        return out
    base = part_name.rsplit("/", 1)[0]
    for rel in rels.findall("rel:Relationship", NS):
        target = rel.attrib.get("Target", "")
        mode = rel.attrib.get("TargetMode", "")
        rel_type = rel.attrib.get("Type", "")
        if mode != "External":
            target = posixpath.normpath(
                target.lstrip("/") if target.startswith("/") else f"{base}/{target}"
            )
        out.append({"type": rel_type, "target": target, "mode": mode, "target_mode": mode or "Internal"})
    return out


def presentation_slide_order(zf: zipfile.ZipFile) -> tuple[list[dict[str, str]], list[str]]:
    errors: list[str] = []
    root = parse_xml(read_text(zf, "ppt/presentation.xml"))
    rels = parse_xml(read_text(zf, "ppt/_rels/presentation.xml.rels"))
    if root is None or rels is None:
        return [], ["presentation slide order graph is missing or invalid"]
    rel_map: dict[str, str] = {}
    for rel in rels.findall("rel:Relationship", NS):
        relationship_id = rel.attrib.get("Id", "")
        target = rel.attrib.get("Target", "")
        if relationship_id and target and rel.attrib.get("TargetMode") != "External":
            rel_map[relationship_id] = posixpath.normpath(
                target.lstrip("/") if target.startswith("/") else f"ppt/{target}"
            )
    ordered: list[dict[str, str]] = []
    seen_ids: set[str] = set()
    for slide_id in root.findall(".//p:sldIdLst/p:sldId", NS):
        stable_id = slide_id.attrib.get("id", "")
        relationship_id = slide_id.attrib.get(f"{{{NS['r']}}}id", "")
        if stable_id in seen_ids:
            errors.append(f"duplicate slide id: {stable_id}")
        seen_ids.add(stable_id)
        part = rel_map.get(relationship_id)
        if not part:
            errors.append(f"slide relationship is unresolved: {relationship_id}")
            continue
        ordered.append({
            "id": stable_id,
            "relationship_id": relationship_id,
            "part": part,
        })
    return ordered, errors


def validate_charts(
    zf: zipfile.ZipFile,
    chart_parts: list[str],
) -> list[str]:
    names = set(zf.namelist())
    errors: list[str] = []
    for chart_part in sorted(set(chart_parts)):
        if chart_part not in names:
            errors.append(f"chart part is missing: {chart_part}")
            continue
        root = parse_xml(read_text(zf, chart_part))
        if root is None:
            errors.append(f"chart XML is invalid: {chart_part}")
            continue
        embedded_workbooks = [
            relationship["target"]
            for relationship in rel_targets(zf, chart_part)
            if relationship.get("target_mode") != "External"
            and relationship["target"].startswith("ppt/embeddings/")
            and relationship["target"].lower().endswith(".xlsx")
        ]
        workbook = None
        if embedded_workbooks:
            try:
                import openpyxl  # type: ignore

                workbook = openpyxl.load_workbook(
                    io.BytesIO(zf.read(embedded_workbooks[0])),
                    data_only=True,
                    read_only=True,
                )
            except (ImportError, OSError, KeyError, ValueError, zipfile.BadZipFile) as error:
                errors.append(
                    f"chart embedded workbook cannot be inspected: {chart_part} -> "
                    f"{embedded_workbooks[0]} ({type(error).__name__})"
                )
        try:
            for series_index, series in enumerate(root.findall(".//c:ser", NS), start=1):
                category_points = series.findall(".//c:cat//c:pt", NS)
                value_points = series.findall(".//c:val//c:pt", NS)
                if category_points and value_points and len(category_points) != len(value_points):
                    errors.append(
                        f"chart cache dimension mismatch: {chart_part} series {series_index} "
                        f"categories={len(category_points)} values={len(value_points)}"
                    )
                if workbook is not None:
                    for role, container in (
                        ("title", series.find("c:tx", NS)),
                        ("categories", series.find("c:cat", NS)),
                        ("values", series.find("c:val", NS)),
                    ):
                        if container is None:
                            continue
                        reference = container.find(".//c:f", NS)
                        cache = next(
                            (
                                item
                                for tag in ("strCache", "numCache")
                                if (item := container.find(f".//c:{tag}", NS)) is not None
                            ),
                            None,
                        )
                        if reference is None or not (reference.text or "").strip() or cache is None:
                            continue
                        try:
                            expected = _chart_formula_values(workbook, reference.text or "")
                        except ValueError as error:
                            errors.append(
                                f"chart source formula is unsupported: {chart_part} series "
                                f"{series_index} {role} ({error})"
                            )
                            continue
                        cached = {
                            int(point.attrib.get("idx", "-1")): (
                                point.find("c:v", NS).text
                                if point.find("c:v", NS) is not None
                                else ""
                            )
                            for point in cache.findall("c:pt", NS)
                        }
                        point_count = cache.find("c:ptCount", NS)
                        declared_count = int(point_count.attrib.get("val", "0")) if point_count is not None else len(cached)
                        if declared_count != len(expected):
                            errors.append(
                                f"chart cache/workbook count mismatch: {chart_part} series "
                                f"{series_index} {role} cache={declared_count} workbook={len(expected)}"
                            )
                            continue
                        for point_index, expected_value in enumerate(expected):
                            actual_value = cached.get(point_index, "")
                            if not _chart_values_equal(actual_value, expected_value):
                                errors.append(
                                    f"chart cache/workbook value mismatch: {chart_part} series "
                                    f"{series_index} {role} point={point_index} "
                                    f"cache={actual_value!r} workbook={expected_value!r}"
                                )
        finally:
            if workbook is not None:
                workbook.close()
        for relationship in rel_targets(zf, chart_part):
            if relationship.get("target_mode") == "External":
                continue
            target = relationship["target"]
            if target not in names:
                errors.append(f"chart dependency is missing: {chart_part} -> {target}")
                continue
            if target.startswith("ppt/embeddings/") and target.lower().endswith(".xlsx"):
                if not zipfile.is_zipfile(io.BytesIO(zf.read(target))):
                    errors.append(f"chart embedded workbook is invalid: {target}")
    return errors


def _chart_formula_values(workbook, formula: str) -> list[object]:
    from openpyxl.utils.cell import range_boundaries  # type: ignore
    from openpyxl.utils.datetime import to_excel  # type: ignore

    normalized = formula.strip().lstrip("=")
    if "[" in normalized or "]" in normalized:
        raise ValueError("external workbook references are not cache-verifiable")
    match = re.fullmatch(
        r"(?:'(?P<quoted>(?:[^']|'')+)'|(?P<plain>[^!]+))!(?P<range>\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?)",
        normalized,
    )
    if match is None:
        raise ValueError(formula)
    sheet_name = (match.group("quoted") or match.group("plain") or "").replace("''", "'")
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"worksheet is missing: {sheet_name}")
    sheet = workbook[sheet_name]
    if not hasattr(sheet, "iter_rows"):
        raise ValueError(f"chart source is not a worksheet: {sheet_name}")
    min_col, min_row, max_col, max_row = range_boundaries(match.group("range").replace("$", ""))
    return [
        to_excel(cell.value, workbook.epoch)
        if isinstance(cell.value, (dt.datetime, dt.date, dt.time))
        else cell.value
        for row in sheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
        )
        for cell in row
    ]


def _chart_values_equal(cached: str | None, expected: object) -> bool:
    cached_text = "" if cached is None else str(cached)
    if expected is None:
        return cached_text == ""
    if isinstance(expected, bool):
        return cached_text.casefold() in ({"true", "1"} if expected else {"false", "0"})
    if isinstance(expected, (int, float)):
        try:
            return abs(float(cached_text) - float(expected)) <= 1e-9
        except ValueError:
            return False
    return cached_text == str(expected)


def slide_text(root) -> str:
    if root is None:
        return ""
    return " ".join(t.text or "" for t in root.findall(".//a:t", NS)).strip()


def shape_inventory(root) -> list[dict[str, object]]:
    """Return stable OOXML identifiers that typed edits can address."""
    if root is None:
        return []
    inventory: list[dict[str, object]] = []
    for kind, xpath in (
        ("shape", ".//p:sp"),
        ("picture", ".//p:pic"),
        ("graphicFrame", ".//p:graphicFrame"),
        ("connector", ".//p:cxnSp"),
        ("group", ".//p:grpSp"),
    ):
        for shape in root.findall(xpath, NS):
            properties = shape.find(".//p:cNvPr", NS)
            if properties is None:
                continue
            text = " ".join(
                item.text or "" for item in shape.findall(".//a:t", NS)
            ).strip()
            placeholder = shape.find(".//p:ph", NS)
            transform = shape.find(".//a:xfrm", NS)
            offset = transform.find("a:off", NS) if transform is not None else None
            extent = transform.find("a:ext", NS) if transform is not None else None
            inventory.append({
                "shapeId": properties.attrib.get("id", ""),
                "shapeName": properties.attrib.get("name", ""),
                "kind": kind,
                "text": text,
                "placeholderType": placeholder.attrib.get("type") if placeholder is not None else None,
                "placeholderIndex": placeholder.attrib.get("idx") if placeholder is not None else None,
                "isTitle": (
                    placeholder is not None
                    and placeholder.attrib.get("type") in {"title", "ctrTitle"}
                ),
                "altText": properties.attrib.get("descr", ""),
                "title": properties.attrib.get("title", ""),
                "bounds": {
                    "x": int(offset.attrib.get("x", "0")) if offset is not None else None,
                    "y": int(offset.attrib.get("y", "0")) if offset is not None else None,
                    "cx": int(extent.attrib.get("cx", "0")) if extent is not None else None,
                    "cy": int(extent.attrib.get("cy", "0")) if extent is not None else None,
                },
            })
    return sorted(
        inventory,
        key=lambda item: (
            int(item["bounds"]["y"] or 0),  # type: ignore[index]
            int(item["bounds"]["x"] or 0),  # type: ignore[index]
            int(str(item["shapeId"]) or 0),
        ),
    )


def nonempty_text_paragraphs(root) -> int:
    if root is None:
        return 0
    count = 0
    for paragraph in root.findall(".//a:p", NS):
        text = " ".join(t.text or "" for t in paragraph.findall(".//a:t", NS)).strip()
        if text:
            count += 1
    return count


def count_placeholders_without_text(root) -> int:
    if root is None:
        return 0
    empty = 0
    for shape in root.findall(".//p:sp", NS):
        if shape.find(".//p:ph", NS) is None:
            continue
        text = " ".join(t.text or "" for t in shape.findall(".//a:t", NS)).strip()
        if not text:
            empty += 1
    return empty


def full_slide_picture_count(root, slide_size: dict[str, int] | None) -> int:
    if root is None or not slide_size:
        return 0
    count = 0
    slide_cx = slide_size.get("cx") or 0
    slide_cy = slide_size.get("cy") or 0
    if not slide_cx or not slide_cy:
        return 0
    for pic in root.findall(".//p:pic", NS):
        xfrm = pic.find(".//a:xfrm", NS)
        if xfrm is None:
            continue
        ext = xfrm.find("a:ext", NS)
        if ext is None:
            continue
        cx = int(ext.attrib.get("cx", "0") or 0)
        cy = int(ext.attrib.get("cy", "0") or 0)
        if cx >= slide_cx * 0.9 and cy >= slide_cy * 0.9:
            count += 1
    return count


def presentation_size(zf: zipfile.ZipFile) -> dict[str, int] | None:
    root = parse_xml(read_text(zf, "ppt/presentation.xml"))
    if root is None:
        return None
    size = root.find("p:sldSz", NS)
    if size is None:
        return None
    return {
        "cx": int(size.attrib.get("cx", "0") or 0),
        "cy": int(size.attrib.get("cy", "0") or 0),
    }


def audit(path: Path) -> dict:
    warnings: list[str] = []
    with zipfile.ZipFile(path) as zf:
        names = set(zf.namelist())
        slide_size = presentation_size(zf)
        physical_slide_names = sorted(
            [name for name in names if re.match(r"ppt/slides/slide\d+\.xml$", name)],
            key=natural_key,
        )
        slide_order, order_errors = presentation_slide_order(zf)
        if slide_order:
            slide_names = [record["part"] for record in slide_order]
        else:
            slide_names = physical_slide_names
            warnings.append("presentation order graph unavailable; using physical slide part order")
        ordered_set = set(slide_names)
        orphan_parts = sorted(set(physical_slide_names) - ordered_set, key=natural_key)
        if orphan_parts:
            order_errors.append("orphan slide parts: " + ", ".join(orphan_parts))
        layouts = [name for name in names if re.match(r"ppt/slideLayouts/slideLayout\d+\.xml$", name)]
        masters = [name for name in names if re.match(r"ppt/slideMasters/slideMaster\d+\.xml$", name)]
        themes = [name for name in names if re.match(r"ppt/theme/theme\d+\.xml$", name)]

        slides = []
        chart_parts: list[str] = []
        comment_details: list[dict[str, object]] = []
        for index, slide_name in enumerate(slide_names, start=1):
            root = parse_xml(read_text(zf, slide_name))
            rels = rel_targets(zf, slide_name)
            text = slide_text(root)
            chart_count = sum(1 for rel in rels if "/charts/" in rel["target"])
            chart_parts.extend(rel["target"] for rel in rels if "/charts/" in rel["target"])
            image_count = sum(1 for rel in rels if "/media/" in rel["target"])
            notes_count = sum(1 for rel in rels if "notesSlide" in rel["type"])
            comment_parts = [
                rel["target"]
                for rel in rels
                if rel["type"].rsplit("/", 1)[-1] == "comments"
                and rel.get("target_mode") != "External"
            ]
            for comment_part in comment_parts:
                comments = parse_xml(read_text(zf, comment_part))
                if comments is None:
                    continue
                for comment in comments:
                    if comment.tag.rsplit("}", 1)[-1] != "cm":
                        continue
                    text_element = next(
                        (item for item in comment if item.tag.rsplit("}", 1)[-1] == "text"),
                        None,
                    )
                    comment_details.append({
                        "slideIndex": index,
                        "slideId": slide_order[index - 1]["id"] if slide_order else None,
                        "part": comment_part,
                        "authorId": comment.attrib.get("authorId", ""),
                        "index": comment.attrib.get("idx", ""),
                        "date": comment.attrib.get("dt", ""),
                        "text": (text_element.text or "") if text_element is not None else "",
                    })
            external_count = sum(1 for rel in rels if rel.get("target_mode") == "External")
            hyperlink_count = sum(1 for rel in rels if "hyperlink" in rel["type"])
            graphic_frames = len(root.findall(".//p:graphicFrame", NS)) if root is not None else 0
            pictures = len(root.findall(".//p:pic", NS)) if root is not None else 0
            shapes = len(root.findall(".//p:sp", NS)) if root is not None else 0
            empty_placeholders = count_placeholders_without_text(root)
            paragraph_count = nonempty_text_paragraphs(root)
            full_slide_pictures = full_slide_picture_count(root, slide_size)
            has_visual_anchor = any([chart_count, image_count, graphic_frames, pictures]) or shapes >= 5
            if index > 1 and not has_visual_anchor:
                warnings.append(f"slide {index} has no visual anchor")
            if index > 1 and not text:
                warnings.append(f"slide {index} has no extractable text")
            if text and len(text) > 900:
                warnings.append(f"slide {index} is text dense ({len(text)} chars)")
            if paragraph_count > 12 and len(text) > 220:
                warnings.append(f"slide {index} has many text paragraphs ({paragraph_count})")
            if empty_placeholders:
                warnings.append(f"slide {index} has {empty_placeholders} empty placeholder(s)")
            if full_slide_pictures and len(text) < 40:
                warnings.append(f"slide {index} may be a low-editability full-slide image")
            slides.append(
                {
                    "index": index,
                    "slide_id": slide_order[index - 1]["id"] if slide_order else None,
                    "part": slide_name,
                    "text": text[:1200],
                    "text_chars": len(text),
                    "shapes": shapes,
                    "pictures": pictures,
                    "graphic_frames": graphic_frames,
                    "image_relationships": image_count,
                    "chart_relationships": chart_count,
                    "notes_relationships": notes_count,
                    "comment_relationships": len(comment_parts),
                    "external_relationships": external_count,
                    "hyperlink_relationships": hyperlink_count,
                    "empty_placeholders": empty_placeholders,
                    "text_paragraphs": paragraph_count,
                    "has_visual_anchor": has_visual_anchor,
                    "full_slide_pictures": full_slide_pictures,
                    "shape_details": shape_inventory(root),
                }
            )

        if len(slides) > 3 and not any(slide["notes_relationships"] for slide in slides):
            warnings.append("deck has no speaker notes")

        chart_errors = validate_charts(zf, chart_parts)
        validation_errors = order_errors + chart_errors
        return {
            "path": str(path),
            "format": "pptx",
            "package_parts": len(names),
            "slides": len(slides),
            "layouts": len(layouts),
            "masters": len(masters),
            "themes": len(themes),
            "slide_size": slide_size,
            "slide_details": slides,
            "slide_order": slide_order,
            "validation_errors": validation_errors,
            "chart_validation_errors": chart_errors,
            "comments": comment_details,
            "orphan_slide_parts": orphan_parts,
            "external_links": sum(slide["external_relationships"] for slide in slides),
            "hyperlinks": sum(slide["hyperlink_relationships"] for slide in slides),
            "warnings": warnings,
        }


def main() -> int:
    parser = argparse.ArgumentParser(description="Audit PPTX OOXML structure.")
    parser.add_argument("--path", required=True, help="Path to a .pptx file")
    parser.add_argument("--pretty", action="store_true", help="Pretty-print JSON")
    args = parser.parse_args()

    path = Path(args.path).expanduser().resolve()
    if not path.exists():
        print(f"File not found: {path}", file=sys.stderr)
        return 3
    if path.suffix.lower() != ".pptx":
        print(f"Expected .pptx file: {path}", file=sys.stderr)
        return 3
    if not zipfile.is_zipfile(path):
        print(f"Not a valid OOXML zip package: {path}", file=sys.stderr)
        return 3

    result = audit(path)
    print(json.dumps(result, ensure_ascii=False, indent=2 if args.pretty else None))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
