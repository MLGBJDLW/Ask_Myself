from __future__ import annotations

import tempfile
import unittest
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import xlsx_structured_editor
import xlsx_audit


class XlsxStructuredEditorTests(unittest.TestCase):
    def _source(self, root: Path) -> Path:
        import openpyxl
        from openpyxl.chart import BarChart, Reference

        path = root / "source.xlsx"
        workbook = openpyxl.Workbook()
        data = workbook.active
        data.title = "Summary"
        data.append(["Metric", "Value"])
        data.append(["Revenue", 100])
        data.append(["Cost", 40])
        inputs = workbook.create_sheet("Inputs")
        inputs["A1"] = "=Summary!B2"
        chart = BarChart()
        chart.add_data(Reference(data, min_col=2, min_row=1, max_row=3), titles_from_data=True)
        chart.set_categories(Reference(data, min_col=1, min_row=2, max_row=3))
        data.add_chart(chart, "D2")
        workbook.save(path)
        return path

    def test_object_level_sheet_name_validation_table_format_and_chart_edits(self) -> None:
        import openpyxl

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = self._source(root)
            output = root / "edited.xlsx"
            result = xlsx_structured_editor.patch_xlsx(source, output, [
                {"op": "rename_sheet", "sheet": "Summary", "newName": "Data"},
                {"op": "set_defined_name", "name": "TotalRef", "formula": "Data!$B$2"},
                {
                    "op": "set_data_validation", "sheet": "Data", "range": "B2:B3",
                    "validationType": "whole", "operator": "between",
                    "formula1": "0", "formula2": "1000", "allowBlank": False,
                },
                {"op": "create_table", "sheet": "Data", "range": "A1:B3", "name": "DataTable"},
                {"op": "set_number_format", "sheet": "Data", "range": "B2:B3", "formatCode": "0.00"},
                {"op": "set_chart_title", "chartPart": "xl/charts/chart1.xml", "title": "Updated metrics"},
            ])

            self.assertIn("xl/workbook.xml", result["changedParts"])
            self.assertIn("xl/styles.xml", result["changedParts"])
            self.assertIn("xl/tables/table1.xml", result["changedParts"])
            self.assertIn("xl/charts/chart1.xml", result["changedParts"])
            workbook = openpyxl.load_workbook(output, data_only=False)
            self.assertIn("Data", workbook.sheetnames)
            self.assertNotIn("Summary", workbook.sheetnames)
            self.assertEqual("=Data!B2", workbook["Inputs"]["A1"].value)
            self.assertEqual("0.00", workbook["Data"]["B2"].number_format)
            self.assertIn("DataTable", workbook["Data"].tables)
            self.assertEqual(1, len(workbook["Data"].data_validations.dataValidation))
            self.assertIn("TotalRef", {item.name for item in workbook.defined_names.values()})
            workbook.close()
            with zipfile.ZipFile(output) as archive:
                chart_xml = archive.read("xl/charts/chart1.xml")
                self.assertIn(b"Updated metrics", chart_xml)
                self.assertIn(b"Data!", chart_xml)
                self.assertNotIn(b"Summary!", chart_xml)

    def test_defined_names_and_sheet_names_reject_external_or_ambiguous_identifiers(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = self._source(root)
            with self.assertRaisesRegex(xlsx_structured_editor.XlsxEditError, "network-closed"):
                xlsx_structured_editor.patch_xlsx(source, root / "external.xlsx", [{
                    "op": "set_defined_name", "name": "Remote", "formula": "[other.xlsx]Sheet1!A1",
                }])
            with self.assertRaisesRegex(xlsx_structured_editor.XlsxEditError, "worksheet name"):
                xlsx_structured_editor.patch_xlsx(source, root / "invalid.xlsx", [{
                    "op": "rename_sheet", "sheet": "Summary", "newName": "bad/name",
                }])

    def test_rename_sheet_rewrites_chart_series_categories_and_headers(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = self._source(root)
            output = root / "chart-renamed.xlsx"
            result = xlsx_structured_editor.patch_xlsx(source, output, [{
                "op": "rename_sheet", "sheet": "Summary", "newName": "Source Data",
            }])
            self.assertIn("xl/charts/chart1.xml", result["changedParts"])
            with zipfile.ZipFile(output) as archive:
                chart_xml = archive.read("xl/charts/chart1.xml")
                self.assertIn(b"'Source Data'!", chart_xml)
                self.assertNotIn(b"Summary!", chart_xml)

    def test_set_chart_data_updates_source_ranges_formulas_and_caches_together(self) -> None:
        import openpyxl

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = self._source(root)
            output = root / "chart-data.xlsx"
            result = xlsx_structured_editor.patch_xlsx(source, output, [{
                "op": "set_chart_data",
                "chartPart": "xl/charts/chart1.xml",
                "seriesIndex": 1,
                "seriesName": "Updated amount",
                "categoryRange": "Summary!$D$10:$D$12",
                "valueRange": "Summary!$E$10:$E$12",
                "categories": ["North", "South", "West"],
                "values": [150, 75.5, 42],
            }])

            self.assertEqual(
                {"xl/charts/chart1.xml", "xl/worksheets/sheet1.xml"},
                set(result["changedParts"]),
            )
            workbook = openpyxl.load_workbook(output, data_only=False)
            sheet = workbook["Summary"]
            self.assertEqual(["North", "South", "West"], [sheet[f"D{row}"].value for row in range(10, 13)])
            self.assertEqual([150, 75.5, 42], [sheet[f"E{row}"].value for row in range(10, 13)])
            self.assertEqual("Updated amount", sheet["B1"].value)
            workbook.close()
            with zipfile.ZipFile(output) as archive:
                chart = ET.fromstring(archive.read("xl/charts/chart1.xml"))
            formulas = [
                item.text or ""
                for item in chart.iter()
                if item.tag.rsplit("}", 1)[-1] == "f"
            ]
            self.assertIn("Summary!$D$10:$D$12", formulas)
            self.assertIn("Summary!$E$10:$E$12", formulas)
            values = [
                item.text or ""
                for item in chart.iter()
                if item.tag.rsplit("}", 1)[-1] == "v"
            ]
            self.assertIn("75.5", values)
            self.assertIn("Updated amount", values)
            self.assertEqual([], xlsx_audit.audit(output)["chart_validation_errors"])

            broken = root / "chart-data-broken.xlsx"
            with zipfile.ZipFile(output) as input_archive, zipfile.ZipFile(broken, "w") as output_archive:
                for info in input_archive.infolist():
                    payload = input_archive.read(info.filename)
                    if info.filename == "xl/charts/chart1.xml":
                        chart = ET.fromstring(payload)
                        cached = chart.find(
                            f".//{{{xlsx_structured_editor.CHART_NS}}}val"
                            f"//{{{xlsx_structured_editor.CHART_NS}}}numCache"
                            f"/{{{xlsx_structured_editor.CHART_NS}}}pt"
                            f"/{{{xlsx_structured_editor.CHART_NS}}}v"
                        )
                        self.assertIsNotNone(cached)
                        cached.text = "999"
                        payload = ET.tostring(chart, encoding="utf-8", xml_declaration=True)
                    output_archive.writestr(info, payload)
            self.assertTrue(
                any(
                    "cache/source value mismatch" in error
                    for error in xlsx_audit.audit(broken)["chart_validation_errors"]
                )
            )

    def test_set_chart_data_rejects_overlapping_or_external_ranges(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = self._source(root)
            base = {
                "op": "set_chart_data",
                "chartPart": "xl/charts/chart1.xml",
                "seriesIndex": 1,
                "categories": ["A", "B"],
                "values": [1, 2],
            }
            with self.assertRaisesRegex(xlsx_structured_editor.XlsxEditError, "overlap"):
                xlsx_structured_editor.patch_xlsx(source, root / "overlap.xlsx", [{
                    **base,
                    "categoryRange": "Summary!A2:A3",
                    "valueRange": "Summary!A2:A3",
                }])
            with self.assertRaisesRegex(xlsx_structured_editor.XlsxEditError, "current workbook"):
                xlsx_structured_editor.patch_xlsx(source, root / "external.xlsx", [{
                    **base,
                    "categoryRange": "[remote.xlsx]Summary!A2:A3",
                }])

    def test_set_range_rejects_ragged_matrix_with_matching_flat_count(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = self._source(root)
            with self.assertRaisesRegex(xlsx_structured_editor.XlsxEditError, "shape"):
                xlsx_structured_editor.patch_xlsx(source, root / "ragged.xlsx", [{
                    "op": "set_range",
                    "sheet": "Summary",
                    "range": "A1:B2",
                    "values": [[1], [2, 3, 4]],
                }])

    def test_sheet_and_chart_operations_preserve_request_order_across_rename(self) -> None:
        import openpyxl

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = self._source(root)
            output = root / "ordered.xlsx"
            xlsx_structured_editor.patch_xlsx(source, output, [
                {"op": "set_value", "sheet": "Summary", "cell": "B2", "value": 999},
                {
                    "op": "set_chart_data",
                    "chartPart": "xl/charts/chart1.xml",
                    "seriesIndex": 1,
                    "categoryRange": "Summary!$D$10:$D$11",
                    "valueRange": "Summary!$E$10:$E$11",
                    "categories": ["North", "South"],
                    "values": [50, 25],
                },
                {"op": "rename_sheet", "sheet": "Summary", "newName": "Data"},
                {"op": "set_value", "sheet": "Data", "cell": "B3", "value": 777},
            ])

            workbook = openpyxl.load_workbook(output, data_only=False)
            self.assertEqual(999, workbook["Data"]["B2"].value)
            self.assertEqual(777, workbook["Data"]["B3"].value)
            self.assertEqual(["North", "South"], [workbook["Data"][f"D{row}"].value for row in (10, 11)])
            self.assertEqual([50, 25], [workbook["Data"][f"E{row}"].value for row in (10, 11)])
            workbook.close()
            with zipfile.ZipFile(output) as archive:
                chart_xml = archive.read("xl/charts/chart1.xml")
            self.assertIn(b"Data!$D$10:$D$11", chart_xml)
            self.assertIn(b"Data!$E$10:$E$11", chart_xml)
            self.assertNotIn(b"Summary!", chart_xml)


if __name__ == "__main__":
    unittest.main()
