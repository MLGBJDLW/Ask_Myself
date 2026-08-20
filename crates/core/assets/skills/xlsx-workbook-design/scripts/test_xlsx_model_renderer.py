#!/usr/bin/env python3
"""Tests for the XLSX model renderer."""

from __future__ import annotations

import json
import tempfile
import unittest
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import xlsx_audit
import xlsx_model_renderer


class XlsxModelRendererTests(unittest.TestCase):
    def test_audit_maps_reordered_sheets_through_workbook_relationships(self) -> None:
        try:
            import openpyxl  # type: ignore
        except ImportError:
            self.skipTest("openpyxl is not installed")
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            original = root / "original.xlsx"
            reordered = root / "reordered.xlsx"
            wb = openpyxl.Workbook()
            wb.active.title = "A"
            wb.active["A1"] = "part-A"
            wb.create_sheet("B")["A1"] = "part-B"
            wb.save(original)
            wb.close()
            with zipfile.ZipFile(original) as source, zipfile.ZipFile(reordered, "w") as target:
                for info in source.infolist():
                    data = source.read(info.filename)
                    if info.filename == "xl/workbook.xml":
                        document = ET.fromstring(data)
                        sheets = document.find(f"{{{xlsx_audit.NS['main']}}}sheets")
                        self.assertIsNotNone(sheets)
                        children = list(sheets)
                        sheets.remove(children[0])
                        sheets.append(children[0])
                        data = ET.tostring(document, encoding="utf-8", xml_declaration=True)
                    target.writestr(info, data)

            report = xlsx_audit.audit(reordered)
            self.assertEqual(["B", "A"], [sheet["name"] for sheet in report["sheet_details"]])
            self.assertTrue(report["sheet_details"][0]["part"].endswith("sheet2.xml"))
            self.assertTrue(report["sheet_details"][1]["part"].endswith("sheet1.xml"))

    def test_create_xlsx_from_spec_writes_tables_formulas_and_qa(self) -> None:
        try:
            import openpyxl  # type: ignore
        except ImportError:
            self.skipTest("openpyxl is not installed")

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp).resolve()
            spec_path = root / "model_spec.json"
            out_path = root / "model.xlsx"
            spec = {
                "title": "Revenue Model",
                "active_sheet": "Summary",
                "qa": {"min_formulas": 4},
                "sheets": [
                    {
                        "name": "Summary",
                        "title": "Revenue Model",
                        "start_cell": "B4",
                        "headers": ["Metric", "2024", "2025", "2026"],
                        "rows": [
                            ["Revenue", 100, 120, 144],
                            ["Gross Margin", 0.42, 0.44, 0.45],
                            ["Gross Profit", None, None, None],
                        ],
                        "table": {"name": "SummaryTable"},
                        "formulas": [
                            {
                                "cell": "C6",
                                "formula": "=C4*C5",
                                "fill_down": {"to_cell": "E6"},
                                "number_format": "$#,##0",
                            },
                            {
                                "cell": "B8",
                                "formula": "=SUM(SummaryTable[2024])",
                                "number_format": "$#,##0",
                            },
                        ],
                        "named_ranges": [{"name": "BaseRevenue", "cell": "C4"}],
                        "validations": [
                            {
                                "range": "C5:E5",
                                "type": "decimal",
                                "operator": "between",
                                "formula1": "0",
                                "formula2": "1",
                            }
                        ],
                        "conditional_formats": [
                            {
                                "range": "C6:E6",
                                "type": "cellIs",
                                "operator": "lessThan",
                                "formula": ["0"],
                                "fill": "FCE4D6",
                            }
                        ],
                        "column_widths": {"B": 20, "C": 14, "D": 14, "E": 14},
                    }
                ],
            }
            spec_path.write_text(json.dumps(spec), encoding="utf-8")

            result = xlsx_model_renderer.create_xlsx_from_spec(
                out_path,
                spec_path,
                workspace_root=root,
            )

            self.assertEqual("pass", result["qa"]["status"])
            self.assertEqual(4, result["qa"]["formulas"])
            self.assertTrue(out_path.exists())
            self.assertTrue(out_path.with_suffix(".xlsx.qa.json").exists())

            wb = openpyxl.load_workbook(out_path, data_only=False)
            try:
                ws = wb["Summary"]
                self.assertEqual("=C4*C5", ws["C6"].value)
                self.assertEqual("=D4*D5", ws["D6"].value)
                self.assertEqual("=E4*E5", ws["E6"].value)
                self.assertEqual("=SUM(SummaryTable[2024])", ws["B8"].value)
                self.assertIn("SummaryTable", ws.tables)
                self.assertIn("BaseRevenue", wb.defined_names)
                self.assertTrue(wb.calculation.fullCalcOnLoad)
            finally:
                wb.close()

    def test_formula_lint_flags_missing_sheet_and_ref_error(self) -> None:
        try:
            import openpyxl  # type: ignore
        except ImportError:
            self.skipTest("openpyxl is not installed")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Model"
        ws["A1"] = "=Missing!A1+#REF!"

        qa = xlsx_model_renderer.audit_workbook_formulas(wb)
        self.assertEqual("fail", qa["status"])
        messages = " ".join(issue["message"] for issue in qa["issues"])
        self.assertIn("missing sheet reference", messages)
        self.assertIn("#REF!", messages)

    def test_rows_keep_leading_equals_literal_and_formula_specs_explicit(self) -> None:
        try:
            import openpyxl  # type: ignore
        except ImportError:
            self.skipTest("openpyxl is not installed")

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp).resolve()
            spec_path = root / "literal_formula_spec.json"
            out_path = root / "literal_formula.xlsx"
            spec_path.write_text(
                json.dumps({
                    "sheets": [{
                        "name": "Inputs",
                        "headers": ["Untrusted text", "Explicit formula"],
                        "rows": [["=WEBSERVICE(\"https://example.invalid/\")", None]],
                        "formulas": [{"cell": "B2", "formula": "=1+1"}],
                    }],
                }),
                encoding="utf-8",
            )

            xlsx_model_renderer.create_xlsx_from_spec(out_path, spec_path, workspace_root=root)
            wb = openpyxl.load_workbook(out_path, data_only=False)
            try:
                ws = wb["Inputs"]
                self.assertEqual("=WEBSERVICE(\"https://example.invalid/\")", ws["A2"].value)
                self.assertEqual("s", ws["A2"].data_type)
                self.assertEqual("=1+1", ws["B2"].value)
                self.assertEqual("f", ws["B2"].data_type)
            finally:
                wb.close()

    def test_formula_inventory_fingerprints_definitions_and_scans_modern_unknown_errors(self) -> None:
        try:
            import openpyxl  # type: ignore
        except ImportError:
            self.skipTest("openpyxl is not installed")
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "evidence.xlsx"
            workbook = openpyxl.Workbook()
            workbook.active.title = "Model"
            workbook.active["A1"] = 1
            workbook.active["A2"] = "=Model!A1+1"
            workbook.active["B1"] = "#SPILL!"
            workbook.active["B1"].data_type = "e"
            workbook.active["B2"] = "#FUTURE!"
            workbook.active["B2"].data_type = "e"
            workbook.save(path)
            workbook.close()

            inventory = xlsx_model_renderer.inspect_formula_inventory(path)
            errors = xlsx_model_renderer.inspect_formula_errors(path)

            self.assertEqual(1, inventory["formulaCells"])
            self.assertEqual(64, len(inventory["fingerprint"]))
            self.assertEqual("Model!A2", inventory["dependencyEdges"][0]["from"])
            self.assertEqual(2, errors["count"])
            by_cell = {item["cell"]: item for item in errors["cells"]}
            self.assertTrue(by_cell["B1"]["known"])
            self.assertFalse(by_cell["B2"]["known"])


if __name__ == "__main__":
    unittest.main()
