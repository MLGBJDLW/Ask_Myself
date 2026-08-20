#!/usr/bin/env python3
"""Skill-owned XLSX model renderer and formula linter.

This renderer deliberately avoids LibreOffice/Excel automation. It writes
workbooks with openpyxl, marks them for automatic recalculation when opened in
Excel, and runs deterministic formula lint checks that can be performed from
the OOXML/workbook structure itself.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import zipfile
from pathlib import Path
from typing import Any, Iterable
from xml.etree import ElementTree as ET


EXCEL_MAX_ROW = 1_048_576
EXCEL_MAX_COL = 16_384
EXCEL_ERRORS = ("#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A")
DEFAULT_TABLE_STYLE = "TableStyleMedium2"
DEFAULT_HEADER_FILL = "1F4E79"
DEFAULT_HEADER_FONT = "FFFFFF"
DEFAULT_INPUT_FONT = "0000FF"
DEFAULT_FORMULA_FONT = "000000"
DEFAULT_LINK_FONT = "008000"

SHEET_REF_RE = re.compile(
    r"(?P<sheet>'(?:[^']|'')+'|[A-Za-z_][A-Za-z0-9_ .]*)!(?P<ref>\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?)"
)
CELL_REF_RE = re.compile(r"(?<![A-Za-z0-9_])(\$?[A-Z]{1,3}\$?\d+)(?::(\$?[A-Z]{1,3}\$?\d+))?(?![A-Za-z0-9_])")
STRUCTURED_REF_RE = re.compile(r"\b([A-Za-z_][A-Za-z0-9_]*)\s*\[")


def _die(message: str, code: int = 1) -> None:
    print(message, file=sys.stderr)
    raise SystemExit(code)


def _missing(pkg: str) -> None:
    print(f"MISSING_DEP: {pkg}", file=sys.stderr)
    print(f"Install with: python -m pip install {pkg}", file=sys.stderr)
    raise SystemExit(2)


def _read_json(path: str) -> dict[str, Any]:
    with Path(path).expanduser().resolve().open("r", encoding="utf-8") as handle:
        data = json.load(handle)
    if not isinstance(data, dict):
        _die("ERROR: workbook spec root must be a JSON object", 3)
    return data


def _write_json(path: Path, data: Any, *, pretty: bool = True) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2 if pretty else None),
        encoding="utf-8",
    )


def _as_list(value: Any) -> list[Any]:
    return value if isinstance(value, list) else []


def _as_dict(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def _as_str(value: Any, default: str = "") -> str:
    return value if isinstance(value, str) else default


def _bool(value: Any, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() not in {"0", "false", "no", "off"}
    return bool(value)


def _clean_hex(value: Any, default: str = "FFFFFF") -> str:
    raw = str(value or default).strip().lstrip("#")
    if len(raw) == 3:
        raw = "".join(ch * 2 for ch in raw)
    if re.fullmatch(r"[0-9a-fA-F]{6}", raw):
        return raw.upper()
    return default.upper()


def _load_openpyxl():
    try:
        import openpyxl  # type: ignore
        from openpyxl import Workbook, load_workbook  # type: ignore
        from openpyxl.chart import AreaChart, BarChart, DoughnutChart, LineChart, PieChart, Reference  # type: ignore
        from openpyxl.comments import Comment  # type: ignore
        from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule  # type: ignore
        from openpyxl.formula.translate import Translator  # type: ignore
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # type: ignore
        from openpyxl.utils import absolute_coordinate, get_column_letter, quote_sheetname, range_boundaries  # type: ignore
        from openpyxl.utils.cell import column_index_from_string, coordinate_to_tuple  # type: ignore
        from openpyxl.workbook.defined_name import DefinedName  # type: ignore
        from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore
        from openpyxl.worksheet.table import Table, TableStyleInfo  # type: ignore
    except ImportError:
        _missing("openpyxl")
    return {
        "openpyxl": openpyxl,
        "Workbook": Workbook,
        "load_workbook": load_workbook,
        "AreaChart": AreaChart,
        "BarChart": BarChart,
        "DoughnutChart": DoughnutChart,
        "LineChart": LineChart,
        "PieChart": PieChart,
        "Reference": Reference,
        "Comment": Comment,
        "CellIsRule": CellIsRule,
        "ColorScaleRule": ColorScaleRule,
        "FormulaRule": FormulaRule,
        "Translator": Translator,
        "Alignment": Alignment,
        "Border": Border,
        "Font": Font,
        "PatternFill": PatternFill,
        "Side": Side,
        "absolute_coordinate": absolute_coordinate,
        "column_index_from_string": column_index_from_string,
        "coordinate_to_tuple": coordinate_to_tuple,
        "DefinedName": DefinedName,
        "DataValidation": DataValidation,
        "Table": Table,
        "TableStyleInfo": TableStyleInfo,
        "get_column_letter": get_column_letter,
        "quote_sheetname": quote_sheetname,
        "range_boundaries": range_boundaries,
    }


def _safe_sheet_name(raw: Any, fallback: str) -> str:
    name = str(raw or fallback).strip()
    name = re.sub(r"[\[\]\:\*\?\/\\]", " ", name)
    name = re.sub(r"\s+", " ", name).strip("' ")
    return (name or fallback)[:31]


def _safe_identifier(raw: Any, fallback: str) -> str:
    name = re.sub(r"[^A-Za-z0-9_]", "", str(raw or "")) or fallback
    if not re.match(r"^[A-Za-z_]", name):
        name = f"_{name}"
    return name[:240]


def _unique_name(base: str, used: set[str], limit: int = 240) -> str:
    name = base[:limit]
    if name not in used:
        used.add(name)
        return name
    for idx in range(2, 10000):
        suffix = str(idx)
        candidate = f"{base[: max(1, limit - len(suffix))]}{suffix}"
        if candidate not in used:
            used.add(candidate)
            return candidate
    raise ValueError(f"cannot create unique name for {base}")


def _cell_tuple(cell: str, libs: dict[str, Any]) -> tuple[int, int]:
    return libs["coordinate_to_tuple"](cell.replace("$", ""))


def _range_from_cells(start_cell: str, rows: int, cols: int, libs: dict[str, Any]) -> str:
    start_row, start_col = _cell_tuple(start_cell, libs)
    end_row = start_row + max(rows, 1) - 1
    end_col = start_col + max(cols, 1) - 1
    return f"{libs['get_column_letter'](start_col)}{start_row}:{libs['get_column_letter'](end_col)}{end_row}"


def _matrix_from_sheet_spec(sheet_spec: dict[str, Any]) -> list[list[Any]]:
    if isinstance(sheet_spec.get("records"), list):
        records = [row for row in sheet_spec["records"] if isinstance(row, dict)]
        headers = _as_list(sheet_spec.get("headers"))
        if not headers and records:
            headers = list(records[0].keys())
        return [headers] + [[record.get(header, "") for header in headers] for record in records]
    if isinstance(sheet_spec.get("headers"), list):
        body = _as_list(sheet_spec.get("rows"))
        rows = [row if isinstance(row, list) else [row] for row in body]
        return [sheet_spec["headers"]] + rows
    return [row if isinstance(row, list) else [row] for row in _as_list(sheet_spec.get("rows"))]


def _write_matrix(ws: Any, start_cell: str, rows: list[list[Any]], libs: dict[str, Any]) -> str | None:
    if not rows:
        return None
    start_row, start_col = _cell_tuple(start_cell, libs)
    max_cols = max(len(row) for row in rows) if rows else 1
    for row_offset, row in enumerate(rows):
        for col_offset in range(max_cols):
            value = row[col_offset] if col_offset < len(row) else None
            cell = ws.cell(row=start_row + row_offset, column=start_col + col_offset, value=value)
            if isinstance(value, str) and value.startswith("="):
                # Matrix/record data is untrusted literal input. Formulas must be
                # expressed through the typed `formulas` collection below.
                cell.data_type = "s"
    return _range_from_cells(start_cell, len(rows), max_cols, libs)


def _apply_header_style(ws: Any, cell_range: str, libs: dict[str, Any], sheet_spec: dict[str, Any]) -> None:
    min_col, min_row, max_col, _ = libs["range_boundaries"](cell_range)
    header_fill = _clean_hex(sheet_spec.get("header_fill"), DEFAULT_HEADER_FILL)
    header_font = _clean_hex(sheet_spec.get("header_font_color"), DEFAULT_HEADER_FONT)
    fill = libs["PatternFill"]("solid", fgColor=header_fill)
    font = libs["Font"](bold=True, color=header_font)
    alignment = libs["Alignment"](horizontal="center", vertical="center", wrap_text=True)
    for col in range(min_col, max_col + 1):
        cell = ws.cell(row=min_row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment


def _apply_sheet_defaults(ws: Any, sheet_spec: dict[str, Any], table_range: str | None, libs: dict[str, Any]) -> None:
    ws.sheet_view.showGridLines = _bool(sheet_spec.get("show_gridlines"), False)
    if sheet_spec.get("tab_color"):
        ws.sheet_properties.tabColor = _clean_hex(sheet_spec.get("tab_color"), "2563EB")
    if sheet_spec.get("freeze_panes") is not None:
        ws.freeze_panes = sheet_spec.get("freeze_panes") or None
    elif table_range:
        min_col, min_row, _, _ = libs["range_boundaries"](table_range)
        ws.freeze_panes = f"{libs['get_column_letter'](min_col)}{min_row + 1}"
    if sheet_spec.get("auto_filter") and table_range:
        ws.auto_filter.ref = table_range
    if sheet_spec.get("orientation"):
        ws.page_setup.orientation = str(sheet_spec["orientation"])
    if sheet_spec.get("print_area"):
        ws.print_area = str(sheet_spec["print_area"])


def _apply_widths(ws: Any, sheet_spec: dict[str, Any], libs: dict[str, Any]) -> None:
    widths = sheet_spec.get("column_widths") or sheet_spec.get("widths") or {}
    if isinstance(widths, list):
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[libs["get_column_letter"](idx)].width = float(width)
    elif isinstance(widths, dict):
        for key, width in widths.items():
            col = str(key).upper()
            if col.isdigit():
                col = libs["get_column_letter"](int(col))
            ws.column_dimensions[col].width = float(width)
    elif sheet_spec.get("auto_width", True):
        for col in range(1, min(ws.max_column, 60) + 1):
            letter = libs["get_column_letter"](col)
            max_len = 0
            for row in range(1, min(ws.max_row, 200) + 1):
                value = ws.cell(row=row, column=col).value
                if value is not None:
                    max_len = max(max_len, len(str(value)))
            if max_len:
                ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 42)


def _style_from_name(cell: Any, role: str, libs: dict[str, Any]) -> None:
    role = role.lower()
    if role in {"input", "assumption", "hardcode"}:
        cell.font = libs["Font"](color=DEFAULT_INPUT_FONT)
        cell.fill = libs["PatternFill"]("solid", fgColor="FFF2CC") if role == "assumption" else cell.fill
    elif role in {"formula", "calculation"}:
        cell.font = libs["Font"](color=DEFAULT_FORMULA_FONT)
    elif role in {"link", "internal_link"}:
        cell.font = libs["Font"](color=DEFAULT_LINK_FONT, underline="single")
    elif role in {"warning", "check"}:
        cell.fill = libs["PatternFill"]("solid", fgColor="FCE4D6")


def _apply_format_to_cell(cell: Any, fmt: dict[str, Any], libs: dict[str, Any]) -> None:
    if fmt.get("role"):
        _style_from_name(cell, str(fmt["role"]), libs)
    if fmt.get("number_format"):
        cell.number_format = str(fmt["number_format"])
    if fmt.get("font_color") or fmt.get("bold") is not None or fmt.get("italic") is not None:
        font_kwargs = {
            "name": fmt.get("font") or cell.font.name,
            "bold": _bool(fmt.get("bold"), bool(cell.font.bold)),
            "italic": _bool(fmt.get("italic"), bool(cell.font.italic)),
            "underline": "single" if _bool(fmt.get("underline"), False) else cell.font.underline,
        }
        if fmt.get("font_color"):
            font_kwargs["color"] = _clean_hex(fmt.get("font_color"), "000000")
        cell.font = libs["Font"](**font_kwargs)
    if fmt.get("fill"):
        cell.fill = libs["PatternFill"]("solid", fgColor=_clean_hex(fmt.get("fill"), "FFFFFF"))
    if fmt.get("align") or fmt.get("valign") or fmt.get("wrap_text") is not None:
        cell.alignment = libs["Alignment"](
            horizontal=fmt.get("align"),
            vertical=fmt.get("valign"),
            wrap_text=_bool(fmt.get("wrap_text"), False),
        )


def _apply_formats(ws: Any, formats: Iterable[Any], libs: dict[str, Any]) -> None:
    for item in formats:
        if not isinstance(item, dict):
            continue
        ref = str(item.get("range") or item.get("cell") or "")
        if not ref:
            continue
        target = ws[ref]
        if hasattr(target, "coordinate"):
            _apply_format_to_cell(target, item, libs)
            continue
        for row in target:
            cells = row if isinstance(row, tuple) else (row,)
            for cell in cells:
                _apply_format_to_cell(cell, item, libs)


def _ensure_formula(value: Any) -> str:
    formula = str(value or "").strip()
    if not formula:
        _die("ERROR: formula cannot be empty", 3)
    return formula if formula.startswith("=") else f"={formula}"


def _formula_has_template(formula: str) -> bool:
    return any(token in formula for token in ("{row}", "{col}", "{col_letter}", "{cell}"))


def _target_cells(item: dict[str, Any], libs: dict[str, Any]) -> list[str]:
    if item.get("range"):
        min_col, min_row, max_col, max_row = libs["range_boundaries"](str(item["range"]))
        return [
            f"{libs['get_column_letter'](col)}{row}"
            for row in range(min_row, max_row + 1)
            for col in range(min_col, max_col + 1)
        ]
    cell = str(item.get("cell") or "").replace("$", "")
    if not cell:
        _die("ERROR: formula item needs cell or range", 3)
    fill = item.get("fill_down")
    if not fill:
        return [cell]
    start_row, start_col = _cell_tuple(cell, libs)
    if isinstance(fill, dict):
        if fill.get("to_cell"):
            end_row, end_col = _cell_tuple(str(fill["to_cell"]), libs)
            return [
                f"{libs['get_column_letter'](col)}{row}"
                for row in range(start_row, end_row + 1)
                for col in range(start_col, end_col + 1)
            ]
        elif fill.get("to_row"):
            end_row = int(fill["to_row"])
        else:
            end_row = start_row + int(fill.get("rows", 1)) - 1
    elif isinstance(fill, int):
        end_row = start_row + fill - 1
    else:
        end_row = int(fill)
    return [f"{libs['get_column_letter'](start_col)}{row}" for row in range(start_row, end_row + 1)]


def _apply_formulas(ws: Any, formulas: Iterable[Any], libs: dict[str, Any]) -> int:
    count = 0
    translator = libs["Translator"]
    for raw in formulas:
        if not isinstance(raw, dict):
            continue
        base_formula = _ensure_formula(raw.get("formula"))
        targets = _target_cells(raw, libs)
        origin = str(raw.get("cell") or targets[0]).replace("$", "")
        for target in targets:
            row, col = _cell_tuple(target, libs)
            if _formula_has_template(base_formula):
                formula = base_formula.format(
                    row=row,
                    col=col,
                    col_letter=libs["get_column_letter"](col),
                    cell=target,
                )
            elif target != origin:
                formula = translator(base_formula, origin=origin).translate_formula(target)
            else:
                formula = base_formula
            cell = ws[target]
            cell.value = formula
            _style_from_name(cell, "formula", libs)
            if raw.get("number_format"):
                cell.number_format = str(raw["number_format"])
            if raw.get("comment"):
                cell.comment = libs["Comment"](str(raw["comment"]), "Nexa")
            count += 1
    return count


def _add_table(ws: Any, table_range: str, sheet_spec: dict[str, Any], used_tables: set[str], libs: dict[str, Any]) -> str | None:
    table_setting = sheet_spec.get("table", True)
    if not table_setting or not table_range:
        return None
    table_spec = table_setting if isinstance(table_setting, dict) else {}
    min_col, min_row, max_col, _ = libs["range_boundaries"](table_spec.get("range") or table_range)
    seen_headers: set[str] = set()
    for col in range(min_col, max_col + 1):
        cell = ws.cell(row=min_row, column=col)
        header = str(cell.value or f"Column{col - min_col + 1}")
        if header in seen_headers:
            base = header
            suffix = 2
            while f"{base}_{suffix}" in seen_headers:
                suffix += 1
            header = f"{base}_{suffix}"
        seen_headers.add(header)
        cell.value = header
    base = _safe_identifier(table_spec.get("name") or sheet_spec.get("table_name") or ws.title, f"{ws.title}Table")
    name = _unique_name(base[:240], used_tables)
    table = libs["Table"](displayName=name, ref=table_spec.get("range") or table_range)
    style = libs["TableStyleInfo"](
        name=table_spec.get("style") or DEFAULT_TABLE_STYLE,
        showFirstColumn=_bool(table_spec.get("show_first_column"), False),
        showLastColumn=_bool(table_spec.get("show_last_column"), False),
        showRowStripes=_bool(table_spec.get("show_row_stripes"), True),
        showColumnStripes=_bool(table_spec.get("show_column_stripes"), False),
    )
    table.tableStyleInfo = style
    ws.add_table(table)
    return name


def _qualified_ref(sheet_name: str, ref: str, libs: dict[str, Any]) -> str:
    ref = str(ref)
    if "!" in ref:
        return ref
    if ":" in ref:
        start, end = ref.split(":", 1)
        coord = f"{libs['absolute_coordinate'](start)}:{libs['absolute_coordinate'](end)}"
    else:
        coord = libs["absolute_coordinate"](ref)
    return f"{libs['quote_sheetname'](sheet_name)}!{coord}"


def _add_named_ranges(wb: Any, sheet_name: str | None, ranges: Iterable[Any], libs: dict[str, Any], used: set[str]) -> int:
    added = 0
    for item in ranges:
        if not isinstance(item, dict):
            continue
        raw_name = item.get("name")
        raw_ref = item.get("ref") or item.get("range") or item.get("cell")
        if not raw_name or not raw_ref:
            continue
        name = _unique_name(_safe_identifier(raw_name, "NamedRange"), used)
        target_sheet = str(item.get("sheet") or sheet_name or "")
        attr_text = str(raw_ref) if "!" in str(raw_ref) else _qualified_ref(target_sheet, str(raw_ref), libs)
        wb.defined_names.add(libs["DefinedName"](name, attr_text=attr_text))
        added += 1
    return added


def _apply_validations(ws: Any, validations: Iterable[Any], libs: dict[str, Any]) -> int:
    count = 0
    for item in validations:
        if not isinstance(item, dict):
            continue
        ref = str(item.get("range") or item.get("cell") or "")
        if not ref:
            continue
        dv = libs["DataValidation"](
            type=item.get("type") or "list",
            operator=item.get("operator"),
            formula1=item.get("formula1"),
            formula2=item.get("formula2"),
            allow_blank=_bool(item.get("allow_blank"), True),
        )
        if item.get("prompt"):
            dv.prompt = str(item["prompt"])
        if item.get("error"):
            dv.error = str(item["error"])
        ws.add_data_validation(dv)
        dv.add(ref)
        count += 1
    return count


def _apply_conditional_formats(ws: Any, formats: Iterable[Any], libs: dict[str, Any]) -> int:
    count = 0
    for item in formats:
        if not isinstance(item, dict):
            continue
        ref = str(item.get("range") or item.get("cell") or "")
        if not ref:
            continue
        kind = str(item.get("type") or "cellIs")
        fill = libs["PatternFill"]("solid", fgColor=_clean_hex(item.get("fill"), "FCE4D6")) if item.get("fill") else None
        font = libs["Font"](color=_clean_hex(item.get("font_color"), "9C0006")) if item.get("font_color") else None
        if kind == "colorScale":
            rule = libs["ColorScaleRule"](
                start_type=item.get("start_type") or "min",
                start_color=_clean_hex(item.get("start_color"), "F8696B"),
                mid_type=item.get("mid_type") or "percentile",
                mid_value=int(item.get("mid_value", 50)),
                mid_color=_clean_hex(item.get("mid_color"), "FFEB84"),
                end_type=item.get("end_type") or "max",
                end_color=_clean_hex(item.get("end_color"), "63BE7B"),
            )
        elif kind == "formula":
            formulas = item.get("formula") if isinstance(item.get("formula"), list) else [item.get("formula")]
            rule = libs["FormulaRule"](formula=[str(f) for f in formulas if f], fill=fill, font=font)
        else:
            formulas = item.get("formula") if isinstance(item.get("formula"), list) else [item.get("formula1") or item.get("formula")]
            rule = libs["CellIsRule"](
                operator=item.get("operator") or "greaterThan",
                formula=[str(f) for f in formulas if f is not None],
                fill=fill,
                font=font,
            )
        ws.conditional_formatting.add(ref, rule)
        count += 1
    return count


def _reference_from_range(ws: Any, ref: str, libs: dict[str, Any]) -> Any:
    min_col, min_row, max_col, max_row = libs["range_boundaries"](ref)
    return libs["Reference"](ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)


def _add_charts(ws: Any, charts: Iterable[Any], libs: dict[str, Any]) -> int:
    chart_types = {
        "area": libs["AreaChart"],
        "bar": libs["BarChart"],
        "column": libs["BarChart"],
        "line": libs["LineChart"],
        "pie": libs["PieChart"],
        "doughnut": libs["DoughnutChart"],
    }
    count = 0
    for item in charts:
        if not isinstance(item, dict):
            continue
        kind = str(item.get("type") or "bar").lower()
        chart = chart_types.get(kind, libs["BarChart"])()
        if kind == "column" and hasattr(chart, "type"):
            chart.type = "col"
        chart.title = item.get("title") or ""
        if item.get("y_axis_title") and hasattr(chart, "y_axis"):
            chart.y_axis.title = str(item["y_axis_title"])
        if item.get("x_axis_title") and hasattr(chart, "x_axis"):
            chart.x_axis.title = str(item["x_axis_title"])
        if item.get("data"):
            data = _reference_from_range(ws, str(item["data"]), libs)
        else:
            min_col = int(item.get("min_col", 2))
            max_col = int(item.get("max_col", min_col))
            min_row = int(item.get("min_row", 1))
            max_row = int(item.get("max_row", ws.max_row))
            data = libs["Reference"](ws, min_col=min_col, max_col=max_col, min_row=min_row, max_row=max_row)
        chart.add_data(data, titles_from_data=_bool(item.get("titles_from_data"), True))
        if item.get("categories"):
            chart.set_categories(_reference_from_range(ws, str(item["categories"]), libs))
        elif item.get("categories_col"):
            min_row = int(item.get("min_row", 1))
            max_row = int(item.get("max_row", ws.max_row))
            cats = libs["Reference"](ws, min_col=int(item["categories_col"]), min_row=min_row + 1, max_row=max_row)
            chart.set_categories(cats)
        chart.height = float(item.get("height", chart.height))
        chart.width = float(item.get("width", chart.width))
        ws.add_chart(chart, item.get("anchor") or "E2")
        count += 1
    return count


def _strip_formula_strings(formula: str) -> str:
    out = []
    in_string = False
    idx = 0
    while idx < len(formula):
        ch = formula[idx]
        if ch == '"':
            in_string = not in_string
            out.append(" ")
        elif in_string:
            out.append(" ")
        else:
            out.append(ch)
        idx += 1
    return "".join(out)


def _unquote_sheet_name(name: str) -> str:
    name = name.strip()
    if name.startswith("'") and name.endswith("'"):
        return name[1:-1].replace("''", "'")
    return name


def _validate_cell_bounds(cell_ref: str, libs: dict[str, Any]) -> bool:
    row, col = _cell_tuple(cell_ref.split(":", 1)[0], libs)
    if not (1 <= row <= EXCEL_MAX_ROW and 1 <= col <= EXCEL_MAX_COL):
        return False
    if ":" in cell_ref:
        end_row, end_col = _cell_tuple(cell_ref.split(":", 1)[1], libs)
        return 1 <= end_row <= EXCEL_MAX_ROW and 1 <= end_col <= EXCEL_MAX_COL
    return True


def _formula_issues_for_cell(
    *,
    formula: str,
    current_sheet: str,
    table_names: set[str],
    defined_names: set[str],
    sheet_names: set[str],
    libs: dict[str, Any],
) -> tuple[list[str], list[str]]:
    issues: list[str] = []
    warnings: list[str] = []
    stripped = _strip_formula_strings(formula.upper())
    if "#REF!" in stripped:
        issues.append("formula contains #REF!")
    for error in EXCEL_ERRORS:
        if error != "#REF!" and error in stripped:
            warnings.append(f"formula text contains {error}")
    if "[" in stripped and re.search(r"\[[^\]]+\]", stripped):
        if re.search(r"\[[^\]]+\][A-Z_]", stripped):
            issues.append("formula contains an external workbook reference")
    if stripped.count("(") != stripped.count(")"):
        issues.append("unbalanced parentheses")
    if ";" in stripped:
        warnings.append("formula uses semicolon separators; OOXML formulas normally use commas")

    found_sheet_spans: list[tuple[int, int]] = []
    for match in SHEET_REF_RE.finditer(stripped):
        found_sheet_spans.append(match.span())
        sheet_name = _unquote_sheet_name(match.group("sheet"))
        ref = match.group("ref")
        if sheet_name not in sheet_names:
            issues.append(f"missing sheet reference: {sheet_name}")
        if not _validate_cell_bounds(ref, libs):
            issues.append(f"cell reference out of Excel bounds: {ref}")

    for match in STRUCTURED_REF_RE.finditer(formula):
        token = match.group(1)
        upper_token = token.upper()
        if token not in table_names and upper_token not in {name.upper() for name in table_names}:
            if token not in defined_names and upper_token not in {name.upper() for name in defined_names}:
                issues.append(f"unknown structured reference table: {token}")

    def inside_sheet_ref(span: tuple[int, int]) -> bool:
        return any(start <= span[0] and span[1] <= end for start, end in found_sheet_spans)

    for match in CELL_REF_RE.finditer(stripped):
        if inside_sheet_ref(match.span()):
            continue
        ref = match.group(0)
        if not _validate_cell_bounds(ref, libs):
            issues.append(f"cell reference out of Excel bounds: {current_sheet}!{ref}")
    return issues, warnings


def audit_workbook_formulas(wb: Any, spec: dict[str, Any] | None = None) -> dict[str, Any]:
    libs = _load_openpyxl()
    spec = spec or {}
    table_names: set[str] = set()
    for ws in wb.worksheets:
        table_names.update(str(name) for name in ws.tables)
    defined_names = set(str(name) for name in wb.defined_names.keys())
    sheet_names = set(wb.sheetnames)
    issues: list[dict[str, str]] = []
    warnings: list[dict[str, str]] = []
    formula_count = 0
    formula_cells: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                    continue
                formula_count += 1
                location = f"{ws.title}!{cell.coordinate}"
                formula_cells.append(location)
                cell_issues, cell_warnings = _formula_issues_for_cell(
                    formula=cell.value,
                    current_sheet=ws.title,
                    table_names=table_names,
                    defined_names=defined_names,
                    sheet_names=sheet_names,
                    libs=libs,
                )
                issues.extend({"cell": location, "message": message} for message in cell_issues)
                warnings.extend({"cell": location, "message": message} for message in cell_warnings)

    min_formulas = int(_as_dict(spec.get("qa")).get("min_formulas") or spec.get("min_formulas") or 0)
    if min_formulas and formula_count < min_formulas:
        issues.append({"cell": "", "message": f"formula count {formula_count} is below required minimum {min_formulas}"})
    status = "fail" if issues else "warn" if warnings else "pass"
    return {
        "kind": "xlsxFormulaQa",
        "status": status,
        "formulas": formula_count,
        "formulaCells": formula_cells[:200],
        "tables": sorted(table_names),
        "namedRanges": sorted(defined_names),
        "issues": issues,
        "warnings": warnings,
        "note": "Formulas are linted without LibreOffice or Excel recalculation; workbook is marked for automatic recalculation when opened.",
    }


def audit_xlsx_formula_integrity(path: str | Path, spec: dict[str, Any] | None = None) -> dict[str, Any]:
    libs = _load_openpyxl()
    wb = libs["load_workbook"](str(Path(path).expanduser().resolve()), data_only=False)
    try:
        result = audit_workbook_formulas(wb, spec=spec)
    finally:
        wb.close()
    result["calculation"] = inspect_formula_cache(path)
    return result


def inspect_formula_cache(path: str | Path) -> dict[str, Any]:
    """Report whether formula cells have non-empty cached values in OOXML."""
    formula_cells = 0
    cached_formula_cells = 0
    with zipfile.ZipFile(Path(path).expanduser().resolve()) as archive:
        worksheet_names = sorted(
            name for name in archive.namelist()
            if name.startswith("xl/worksheets/") and name.endswith(".xml")
        )
        for name in worksheet_names:
            root = ET.fromstring(archive.read(name))
            for cell in root.iter():
                if cell.tag.rsplit("}", 1)[-1] != "c":
                    continue
                formula = next(
                    (child for child in cell if child.tag.rsplit("}", 1)[-1] == "f"),
                    None,
                )
                if formula is None:
                    continue
                formula_cells += 1
                cached = next(
                    (child for child in cell if child.tag.rsplit("}", 1)[-1] == "v"),
                    None,
                )
                if cached is not None and cached.text not in {None, ""}:
                    cached_formula_cells += 1
    if formula_cells == 0:
        level = "not_applicable"
    elif cached_formula_cells == formula_cells:
        level = "cached_values_present"
    elif cached_formula_cells == 0:
        level = "not_calculated"
    else:
        level = "partially_cached"
    return {
        "level": level,
        "formulaCells": formula_cells,
        "cachedFormulaCells": cached_formula_cells,
        "coverage": 1.0 if formula_cells == 0 else cached_formula_cells / formula_cells,
        "nativeRecalculationProven": False,
    }


def _build_workbook(spec: dict[str, Any]) -> tuple[Any, dict[str, Any]]:
    libs = _load_openpyxl()
    wb = libs["Workbook"]()
    wb.remove(wb.active)
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.properties.creator = _as_str(spec.get("creator"), "Nexa")
    wb.properties.title = _as_str(spec.get("title"), "Workbook")

    sheets = spec.get("sheets") or []
    if not isinstance(sheets, list) or not sheets:
        _die("ERROR: create_xlsx spec requires non-empty 'sheets' array", 3)

    used_sheet_names: set[str] = set()
    used_table_names: set[str] = set()
    used_defined_names: set[str] = set(str(name) for name in wb.defined_names.keys())
    metrics = {
        "sheets": 0,
        "tables": 0,
        "formulas": 0,
        "charts": 0,
        "validations": 0,
        "conditionalFormats": 0,
        "namedRanges": 0,
    }

    for index, sheet_spec_raw in enumerate(sheets, start=1):
        if not isinstance(sheet_spec_raw, dict):
            _die(f"ERROR: sheet {index} must be an object", 3)
        sheet_spec = sheet_spec_raw
        name = _unique_name(_safe_sheet_name(sheet_spec.get("name"), f"Sheet{index}"), used_sheet_names, 31)
        ws = wb.create_sheet(name)
        metrics["sheets"] += 1

        if sheet_spec.get("title"):
            title_cell = str(sheet_spec.get("title_cell") or "B2")
            ws[title_cell] = str(sheet_spec["title"])
            ws[title_cell].font = libs["Font"](bold=True, size=16, color="111827")
        if sheet_spec.get("subtitle"):
            subtitle_cell = str(sheet_spec.get("subtitle_cell") or "B3")
            ws[subtitle_cell] = str(sheet_spec["subtitle"])
            ws[subtitle_cell].font = libs["Font"](color="64748B")

        rows = _matrix_from_sheet_spec(sheet_spec)
        start_cell = str(sheet_spec.get("start_cell") or sheet_spec.get("data_start_cell") or "A1")
        table_range = _write_matrix(ws, start_cell, rows, libs)
        if table_range and sheet_spec.get("header_style", True):
            _apply_header_style(ws, table_range, libs, sheet_spec)
        table_name = _add_table(ws, table_range, sheet_spec, used_table_names, libs) if rows else None
        if table_name:
            metrics["tables"] += 1

        metrics["formulas"] += _apply_formulas(ws, _as_list(sheet_spec.get("formulas")), libs)
        metrics["namedRanges"] += _add_named_ranges(wb, name, _as_list(sheet_spec.get("named_ranges")), libs, used_defined_names)
        metrics["validations"] += _apply_validations(ws, _as_list(sheet_spec.get("validations")), libs)
        metrics["conditionalFormats"] += _apply_conditional_formats(ws, _as_list(sheet_spec.get("conditional_formats")), libs)
        _apply_formats(ws, _as_list(sheet_spec.get("formats")), libs)
        _apply_widths(ws, sheet_spec, libs)
        _apply_sheet_defaults(ws, sheet_spec, table_range, libs)
        metrics["charts"] += _add_charts(ws, _as_list(sheet_spec.get("charts")), libs)

    metrics["namedRanges"] += _add_named_ranges(wb, None, _as_list(spec.get("named_ranges")), libs, used_defined_names)
    active = spec.get("active_sheet")
    if active and active in wb.sheetnames:
        wb.active = wb.sheetnames.index(str(active))
    return wb, metrics


def create_xlsx_from_spec(
    path: str | Path,
    spec_path: str | Path,
    workspace_root: str | Path | None = None,
    *,
    pretty: bool = True,
) -> dict[str, Any]:
    root = Path(workspace_root).resolve() if workspace_root else Path.cwd().resolve()
    out_path = Path(path).expanduser().resolve()
    try:
        out_path.relative_to(root)
    except ValueError:
        _die(f"ERROR: output path escapes workspace: {out_path}", 3)
    if out_path.suffix.lower() != ".xlsx":
        _die("ERROR: output path must end with .xlsx", 3)
    spec_path = Path(spec_path).expanduser().resolve()
    try:
        spec_path.relative_to(root)
    except ValueError:
        _die(f"ERROR: spec path escapes workspace: {spec_path}", 3)
    spec = _read_json(str(spec_path))
    wb, metrics = _build_workbook(spec)
    qa = audit_workbook_formulas(wb, spec=spec)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    qa["calculation"] = inspect_formula_cache(out_path)
    qa_path = out_path.with_suffix(".xlsx.qa.json")
    result = {
        "kind": "xlsxModelRender",
        "path": str(out_path),
        "specPath": str(spec_path),
        "metrics": metrics,
        "qa": qa,
        "qaPath": str(qa_path),
    }
    _write_json(qa_path, result, pretty=pretty)
    return result


def main() -> int:
    parser = argparse.ArgumentParser(description="Create XLSX workbooks from a structured JSON model spec.")
    parser.add_argument("--path", required=True, help="Output .xlsx path under the workspace")
    parser.add_argument("--spec", required=True, help="Workbook JSON spec path under the workspace")
    parser.add_argument("--lint-only", action="store_true", help="Lint formulas in an existing workbook instead of creating one")
    parser.add_argument("--compact", action="store_true", help="Write compact JSON output")
    args = parser.parse_args()

    if args.lint_only:
        result = audit_xlsx_formula_integrity(args.path)
    else:
        result = create_xlsx_from_spec(
            args.path,
            args.spec,
            workspace_root=Path.cwd(),
            pretty=not args.compact,
        )
    print(json.dumps(result, ensure_ascii=False, indent=None if args.compact else 2))
    status = result.get("status") or _as_dict(result.get("qa")).get("status")
    return 0 if status != "fail" else 4


if __name__ == "__main__":
    raise SystemExit(main())
